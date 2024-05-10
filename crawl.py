import openpyxl
from selenium import webdriver

cookies_object_list = []
try:
  wb_cookie = openpyxl.load_workbook('cookies.xlsx')
  ws = wb_cookie.active
  for row in ws.iter_rows():
    cookies_object_list.append({"name": row[0].value, "value": row[1].value})
except FileNotFoundError:
  print("Error: File 'cookies.xlsx' not found.")
finally:
  wb_cookie.close()

wb_comment = openpyxl.Workbook()
sheet_default = wb_comment.active
sheet_default.cell(row=1, column=1).value = "Tên"
sheet_default.cell(row=1, column=2).value = "Số lượng"
row_sheet_default_index = 2

driver = webdriver.Firefox(executable_path="./geckodriver/geckodriver.exe")
for cookies_object in cookies_object_list:
    cookies = []
    id = ""
    for item in cookies_object["value"].split("; "):
        name, value = item.split("=")
        if name=="c_user":
            id = value
        cookies.append({"name": name, "value": value})
    driver.get("http://www.facebook.com/{id}/allactivity?activity_history=false&category_key=COMMENTSCLUSTER&manage_mode=false&should_load_landing_page=false")

    # add cookie to browser
    for cookie in cookies:
        driver.add_cookie(cookie)
    
    # refresh and wait
    driver.refresh()

    # get comment_tag_html
    comment_history_tag = driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[4]")
    comments_tag = comment_history_tag.find_elements_by_xpath(".//div/div[2]/div[1]/div/a/div[1]/div[2]/div/div/div")
    comment_counter = 0

    comments = {}
    for comment_tag in comments_tag:
        location = ""
        content = ""

        try :
            location_tag = comment_tag.find_element_by_xpath(".//div[1]/span/span/span/div/strong[3]/object/a")
            location = location_tag.text
        except:
            location = "Đến chính mình"
        
        content_tag = comment_tag.find_element_by_xpath(".//div[2]/span/span")
        content = content_tag.text

        if location not in comments:
            comments[location] = {}
        if content not in comments[location]:
           comments[location][content] = 0
        comments[location][content] += 1
        comment_counter += 1

    # save to defaut sheet commet number
    sheet_default.cell(row=row_sheet_default_index,column=1).value = cookies_object["name"]
    sheet_default.cell(row=row_sheet_default_index,column=2).value = comment_counter
    row_sheet_default_index += 1
    # create sheet "name" to save comments
    ws = wb_comment.create_sheet(cookies_object["name"])
    ws.cell(row=1, column=1).value = "Nơi comment"
    ws.cell(row=1, column=2).value = "Nội dung"
    ws.cell(row=1, column=3).value = "Số lượng"
    row_index = 2
    for location, contents in comments.items():
        ws.cell(row=row_index, column=1).value = location
        for content, len in contents.items():
            ws.cell(row=row_index, column=2).value = content
            ws.cell(row=row_index, column=3).value = len
            row_index += 1  # Move to the next row for the next comment

# Save the workbook as an Excel file
wb_comment.save("comments.xlsx")
print("Comments saved to comments.xlsx")
driver.close()
