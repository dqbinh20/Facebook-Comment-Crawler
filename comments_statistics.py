from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import requests
from datetime import datetime
from tkinter import filedialog
from typing import List, Dict

# Constants
WEB_DRIVER_WAIT_TIME = 3
SLEEP_INTERVAL = 2

# Function to load cookies from an Excel file
def load_cookies_file() -> List[Dict[str, str]]:
    """
    Loads cookies from an Excel file.

    Returns:
        A list of cookie objects, each containing 'name' and 'value' keys.
    """
    file_path = filedialog.askopenfilename(title="Chọn file cookies.xlsx")
    cookies_object_list: List[Dict[str, str]] = []

    if not file_path:
        print("No file selected.")
        return None

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows():
            cookies_object_list.append({"name": row[0].value, "value": str(row[1].value)})
    except FileNotFoundError:
        print("Error: File 'cookies.xlsx' not found.")
        return None
    except Exception as e:
        print(f"An error occurred while loading cookies: {e}")
        return None

    return cookies_object_list

# Helper function to handle cookie parsing
def parse_cookies(cookies_object: Dict[str, str]) -> tuple[List[Dict[str, str]], str]:
    """
    Parses cookies from a single cookie object.

    Args:
        cookies_object: A dictionary containing 'name' and 'value' keys.

    Returns:
        A tuple containing a list of parsed cookie dictionaries and the user ID.
    """
    cookies: List[Dict[str, str]] = []
    user_id: str = ""
    try:
        for item in cookies_object["value"].split("; "):
            name, value = item.split("=")
            if name == "c_user":
                user_id = value
            cookies.append({"name": name, "value": value})
    except ValueError as e:
        print(f"Error parsing cookies for {cookies_object['name']}: {e}")
    return cookies, user_id

# Function to crawl comments using the provided cookies
def crawl_comment(cookies_object_list: List[Dict[str, str]]) -> openpyxl.Workbook:
    """
    Crawls comments from Facebook using the provided cookies.

    Args:
        cookies_object_list: A list of cookie objects.

    Returns:
        An Excel workbook containing the crawled comments.
    """
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Tên", "Số lượng"])

    driver = webdriver.Chrome()
    driver.get("http://www.facebook.com/")

    group_facebook_name: Dict[str, str] = {}

    for cookies_object in cookies_object_list:
        cookies, user_id = parse_cookies(cookies_object)
        if not cookies:
            continue

        for cookie in cookies:
            driver.add_cookie(cookie)

        try:
            driver.get(f"http://www.facebook.com/{user_id}/allactivity?activity_history=false&category_key=COMMENTSCLUSTER&manage_mode=false&should_load_landing_page=false")
            comment_history_tag = WebDriverWait(driver, WEB_DRIVER_WAIT_TIME).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[4]"))
            )
            comments_tag = comment_history_tag.find_elements(By.XPATH,"div/div[2]/div[1]/div/a")
            previous_comment_count = len(comments_tag)
        except TimeoutException:
            print(f"Cookie error for {cookies_object['name']}")
            continue

        try:
            date_comment = driver.find_element(By.XPATH,"/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[4]/div[1]/div/div/div/div/div/h2/span/span")
            dmy = date_comment.text.split(" ")
            if (int(datetime.now().strftime("%d")) != int(dmy[0]) or 
                int(datetime.now().strftime("%m")) != int(dmy[2].replace(",", "")) or 
                int(datetime.now().strftime("%Y")) != int(dmy[3])):
                print(f"{cookies_object['name']} hasn't commented today")
                summary_sheet.append([cookies_object["name"], 0])
                continue
        except NoSuchElementException:
            print(f"Unable to find comment date for {cookies_object['name']}")
            continue

        while True:
            time.sleep(SLEEP_INTERVAL)
            comment_history_tag = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[4]")
            comments_tag = comment_history_tag.find_elements(By.XPATH,"div/div[2]/div[1]/div/a")
            if len(comments_tag) == previous_comment_count:
                break
            previous_comment_count = len(comments_tag)

        comments: Dict[str, Dict[str, int]] = {}
        link_list: List[List[str]] = []
        contents: List[str] = []

        for tag in comments_tag:
            contents.append(tag.text.split("\n")[1])
            link_list.append(tag.get_attribute("href").split("/"))

        for i, link in enumerate(link_list):
            location = link[3]

            if location == "groups":
                if link[4] not in group_facebook_name:
                    response_group = requests.get(f"https://www.facebook.com/groups/{link[4]}")
                    group_facebook_name[link[4]] = response_group.text.split('<title>')[1].split("</title>")[0]
                location = group_facebook_name[link[4]]

            if location not in comments:
                comments[location] = {}

            if contents[i] not in comments[location]:
                comments[location][contents[i]] = 0

            comments[location][contents[i]] += 1

        summary_sheet.append([cookies_object["name"], len(contents)])

        user_sheet = workbook.create_sheet(cookies_object["name"])
        user_sheet.append(["Nơi comment", "Nội dung", "Số lượng"])

        for location, content_dict in comments.items():
            for content, count in content_dict.items():
                user_sheet.append([location, content, count])

    driver.quit()
    return workbook

# Function to save the comments workbook
def save_wb_comment_file(workbook: openpyxl.Workbook):
    """
    Saves the comments workbook to a file.

    Args:
        workbook: The Excel workbook containing the comments.
    """
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="comments.xlsx", filetypes=[("Excel file", "*.xlsx")])
    if file_path:
        try:
            workbook.save(file_path)
            print("File saved successfully!")
        except Exception as e:
            print(f"An error occurred while saving the file: {e}")

if __name__ == "__main__":
    cookies_object_list = load_cookies_file()
    if cookies_object_list:
        wb_comment = crawl_comment(cookies_object_list)
        save_wb_comment_file(wb_comment)