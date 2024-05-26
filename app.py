from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import requests
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from tkinter import filedialog
from typing import List, Dict
import json

# Constants
WEB_DRIVER_WAIT_TIME = 3
SLEEP_INTERVAL = 2

LINK_COMMENT_HISTORY_PAGE = "https://www.facebook.com/{id}/allactivity?activity_history=false&category_key=COMMENTSCLUSTER&manage_mode=false&should_load_landing_page=false" 
LINK_POST_AND_COMMENT_IN_GROUP_PAGE = "https://www.facebook.com/{id}/allactivity?activity_history=false&category_key=GROUPPOSTS&manage_mode=false&should_load_landing_page=false"

class DomElementPaths:
    def __init__(self, link_page):
        if (link_page == LINK_COMMENT_HISTORY_PAGE):
            self.FULL_XPATH_COMMENT_HISTORY_LATEST = "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[4]"
            self.XPATH_COMMENT_LINKS = "div/div[2]/div[1]/div/a"
            self.FULL_XPATH_COMMENT_DATE_LATEST  = "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[4]/div[1]/div/div/div/div/div/h2/span/span"
        elif (link_page == LINK_POST_AND_COMMENT_IN_GROUP_PAGE):
            self.FULL_XPATH_COMMENT_HISTORY_LATEST = "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[2]"
            self.XPATH_COMMENT_LINKS = "div/div/div[1]/div/a"
            self.FULL_XPATH_COMMENT_DATE_LATEST  = "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[2]/div[1]/div/div/div/div/div/h2/span/span"

# Function to load cookies from an Excel file
def load_cookies_file() -> List[Dict[str, str]]:
    """
    Loads cookies from an Excel file.

    Returns:
        A list of cookie objects, each containing 'name' and 'value' keys.
    """
    file_path = filedialog.askopenfilename(title="Chọn file cookies.xlsx")
    cookies_object_list: List[Dict[str, str]] = []

    if not file_path:
        print("No file selected.")
        return None

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows():
            cookies_object_list.append({"name": row[0].value, "value": str(row[1].value)})
    except FileNotFoundError:
        print("Error: File 'cookies.xlsx' not found.")
        return None
    except Exception as e:
        print(f"An error occurred while loading cookies: {e}")
        return None

    return cookies_object_list

# Helper function to handle cookie parsing
def parse_cookies(cookies_object: Dict[str, str]) -> tuple[List[Dict[str, str]], str]:
    """
    Parses cookies from a single cookie object.

    Args:
        cookies_object: A dictionary containing 'name' and 'value' keys.

    Returns:
        A tuple containing a list of parsed cookie dictionaries and the user ID.
    """
    cookies: List[Dict[str, str]] = []
    user_id: str = ""
    try:
        for item in cookies_object["value"].split("; "):
            name, value = item.split("=")
            if name == "c_user":
                user_id = value
            cookies.append({"name": name, "value": value})
    except ValueError as e:
        print(f"Error parsing cookies for {cookies_object['name']}: {e}")
    return cookies, user_id

# Function to crawl comments using the provided cookies
def crawl_comment(cookies_object_list: List[Dict[str, str]]) -> openpyxl.Workbook:
    """
    Crawls comments from Facebook using the provided cookies.

    Args:
        cookies_object_list: A list of cookie objects.

    Returns:
        An Excel workbook containing the crawled comments.
    """
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Tên", "Số lượng"])

    chrome_options = Options()
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=chrome_options)
    driver.get("http://www.facebook.com/")

    group_facebook_name: Dict[str, str] = {}
    try:    
        with open('group_facebook_name.json', 'r', encoding='utf-8') as file:
            group_facebook_name = json.load(file)
    except: 
        pass

    for cookies_object in cookies_object_list:
        cookies, user_id = parse_cookies(cookies_object)
        if not cookies:
            continue

        driver.delete_all_cookies()
        for cookie in cookies:
            driver.add_cookie(cookie)
        
        try:
            driver.get(LINK_COMMENT_HISTORY_PAGE)
            paths = DomElementPaths(LINK_COMMENT_HISTORY_PAGE)
            comment_history_tag = WebDriverWait(driver, WEB_DRIVER_WAIT_TIME).until(
                EC.presence_of_element_located((By.XPATH, paths.FULL_XPATH_COMMENT_HISTORY_LATEST))
            )
            comments_tag = comment_history_tag.find_elements(By.XPATH,paths.XPATH_COMMENT_LINKS)
            previous_comment_count = len(comments_tag)
        except TimeoutException:
            try:
                driver.get(LINK_POST_AND_COMMENT_IN_GROUP_PAGE)
                paths = DomElementPaths(LINK_POST_AND_COMMENT_IN_GROUP_PAGE)
                comment_history_tag = WebDriverWait(driver, WEB_DRIVER_WAIT_TIME).until(
                    EC.presence_of_element_located((By.XPATH, paths.FULL_XPATH_COMMENT_HISTORY_LATEST))
                )
                comments_tag = comment_history_tag.find_elements(By.XPATH,paths.XPATH_COMMENT_LINKS)
                previous_comment_count = len(comments_tag)
            except TimeoutException:
                print(f"Cookie error for {cookies_object['name']}")
                summary_sheet.append([cookies_object["name"], "Cookie error"])
                continue

        # check comment today
        try:
            date_comment = driver.find_element(By.XPATH,paths.FULL_XPATH_COMMENT_DATE_LATEST)
            dmy = date_comment.text.split(" ")
            if (int(datetime.now().strftime("%d")) != int(dmy[0]) or 
                int(datetime.now().strftime("%m")) != int(dmy[2].replace(",", "")) or 
                int(datetime.now().strftime("%Y")) != int(dmy[3])):
                print(f"{cookies_object['name']} hasn't commented today")
                summary_sheet.append([cookies_object["name"], 0])
                continue
        except NoSuchElementException:
            print(f"Unable to find comment date for {cookies_object['name']}")
            continue
        
        # check if notice facebook appear
        try:
            time.sleep(SLEEP_INTERVAL)
            driver.find_element(By.XPATH, '//div[@aria-label="Đóng"]').click()
            print(f"{cookies_object["name"]} - Closed notice")
        except:
            pass

        # Scroll to end comment list latest
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(SLEEP_INTERVAL)
            comment_history_tag = driver.find_element(By.XPATH, paths.FULL_XPATH_COMMENT_HISTORY_LATEST)
            comments_tag = comment_history_tag.find_elements(By.XPATH,paths.XPATH_COMMENT_LINKS)
            if len(comments_tag) == previous_comment_count:
                break
            previous_comment_count = len(comments_tag)

        # 
        comments_tag_text_link = [] 
        for item in comments_tag:
            comments_tag_text_link.append((item.text,item.get_attribute("href")))

        comments: Dict[str, Dict[str, Dict[str, int]]] = {}
        link_list: List[List[str]] = []
        contents: List[str] = []
        authors : List[str] = []
        # find user_name
        user_name = ""
        try:
            user_name = comments_tag[0].text[:comments_tag[0].text.find(" đã ")]
        except:
            pass

        # loop to check comment state
        cmt_wait = 0
        cmt_deny = 0
        for tag in comments_tag_text_link:

            # type comment
            aria_label = ""
            if "đã bình luận" in tag[0]:
                aria_label = f'//*[starts-with(@aria-label, "Bình luận dưới tên {user_name}")]'
            elif "đã phản hồi" in tag[0]:
                aria_label = f'//*[starts-with(@aria-label, "Phản hồi bình luận của")]'
            elif "đã trả lời" in tag[0]:
                aria_label = f'//*[starts-with(@aria-label, "Phản hồi bình luận của {user_name} dưới tên {user_name}")]'

            driver.get(tag[1])  

            # if reel, click button cmt to show cmts
            if (tag[1].split("/")[3]=="reel"):
                try:
                    cmtButton = WebDriverWait(driver, WEB_DRIVER_WAIT_TIME).until(EC.presence_of_element_located((By.XPATH, '//*[@aria-label="Bình luận"]')))
                    cmtButton.click()
                except TimeoutException:
                    print("Không tìm thấy nút bình luận trong reel")
                    continue
            
            # check comment's state
            try:
                comment_in_group = WebDriverWait(driver, WEB_DRIVER_WAIT_TIME).until(EC.presence_of_element_located((By.XPATH, aria_label)))
                text_comment_in_group = comment_in_group.text

                if text_comment_in_group.find("\nĐang chờ\nTìm hiểu thêm") != -1:
                    print(f"{user_name} có một comment đang chờ duyệt")
                    cmt_wait += 1
                    continue
                if text_comment_in_group.find("\nBị từ chối") != -1:
                    print(f"{user_name} có một comment đã bị từ chối")
                    cmt_deny +=1
                    continue
            except NoSuchElementException:
                if tag[1].split("/")[3]=="reel":
                    pass
                else:
                    print("Không tìm thấy cmt")
                    continue
            except TimeoutException:
                if tag[1].split("/")[3]=="reel":
                    pass
                else:
                    try:
                        # 
                        comment_in_group = WebDriverWait(driver, WEB_DRIVER_WAIT_TIME).until(EC.presence_of_element_located((By.XPATH, f'//*[starts-with(@aria-label, "{user_name} đáp lại phản hồi")]')))
                        text_comment_in_group = comment_in_group.text
                        if text_comment_in_group.find("\nĐang chờ\nTìm hiểu thêm") != -1:
                            print(f"{user_name} có một comment đang chờ duyệt")
                            cmt_wait += 1
                            continue
                        if text_comment_in_group.find("\nBị từ chối") != -1:
                            print(f"{user_name} có một comment đã bị từ chối")
                            cmt_deny +=1
                            continue
                    except:
                        print("Không tìm thấy cmt")
                        continue
            except Exception as e:
                print(str(e))

            # if the comment is approved
            print(f"{user_name} có một comment đã được chấp nhận")
            # tìm tên người viết bài
            author = tag[0][tag[0].find("của ")+3:tag[0].find("\n")]
            try:
                author = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div/div/div/div/div/div/div/div/div/div/div/div[13]/div/div/div[2]/div/div[2]/div/div[1]/span/h2/span/a/strong/span").text
            except NoSuchElementException:
                try:
                    author = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div[4]/div/div/div/div/div/div/div/div[1]/div/div/div/div/div/div/div/div/div/div/div[13]/div/div/div[2]/div/div[2]/div/div[1]/span/h2/span/span/a/strong/span").text
                except NoSuchElementException:
                    try:
                        author = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div[1]/div/div[2]/div/div/div/div[1]/div/div/div[2]/div[1]/div/div[1]/div[3]/div/div/div[1]/div/div/div[2]/div/div[1]/div/h2/span/span/span").text
                    except NoSuchElementException:
                        try:
                            author = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div[2]/div/div/div[4]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[13]/div/div/div[2]/div/div[2]/div/div[1]/span/h3/span/span/a/strong/span").text
                        except Exception:
                            pass
                    except Exception:
                        pass
                except Exception:
                    pass
            except Exception:
                pass

            contents.append(tag[0].split("\n")[1])
            link_list.append(tag[1].split("/"))
            authors.append(author)

        print(f"{user_name} có tổng cộng {len(comments_tag_text_link)}, từ chối {cmt_deny}, chờ duyệt {cmt_wait}")

        print(f"{cookies_object['name']} có {len(contents)}")

        for i, link in enumerate(link_list):
            
            # tìm tên group
            location = link[3]
            if location == "groups":
                try:
                    if link[4] not in group_facebook_name:
                        response_group = requests.get(f"https://www.facebook.com/groups/{link[4]}")
                        group_facebook_name[link[4]] = response_group.text.split('<title>')[1].split("</title>")[0]
                except:
                    group_facebook_name[link[4]] = link[4]
                location = group_facebook_name[link[4]]
            
            ###### thống kê ##########
            if location not in comments:
                comments[location] = {}

            if contents[i] not in comments[location]:
                comments[location][contents[i]] = {}
            
            if authors[i] not in comments[location][contents[i]]:
                comments[location][contents[i]][authors[i]] = 0

            comments[location][contents[i]][authors[i]] += 1

        summary_sheet.append([cookies_object["name"], len(contents)])

        user_sheet = workbook.create_sheet(cookies_object["name"])
        user_sheet.append(["Nơi comment", "Nội dung", "Tác giả bài viết", "Số lượng"])

        for location, content_author_dict in comments.items():
            for content, author_dict in content_author_dict.items():
                for author, count in author_dict.items():
                    user_sheet.append([location, content, author, count])

    try:
        with open('group_facebook_name.json', 'w', encoding='utf-8') as file:
            json.dump(group_facebook_name, file, ensure_ascii=False, indent=4)
    except:
        pass
    driver.quit()
    return workbook

# Function to save the comments workbook
def save_wb_comment_file(workbook: openpyxl.Workbook):
    """
    Saves the comments workbook to a file.

    Args:
        workbook: The Excel workbook containing the comments.
    """
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="comments.xlsx", filetypes=[("Excel file", "*.xlsx")])
    if file_path:
        try:
            workbook.save(file_path)
            print("File saved successfully!")
        except Exception as e:
            print(f"An error occurred while saving the file: {e}")

if __name__ == "__main__":
    cookies_object_list = load_cookies_file()
    if cookies_object_list:
        wb_comment = crawl_comment(cookies_object_list)
        save_wb_comment_file(wb_comment)